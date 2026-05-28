import csv
import json
from collections import deque
from datetime import datetime
from pathlib import Path
from queue import Empty, Queue
from threading import Lock

from flask import Response, jsonify, request, send_from_directory
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


DATA_DIR = Path("data")
HISTORY_FILE = DATA_DIR / "events_history.json"
MAX_HISTORY_SIZE = 1000

history = deque(maxlen=MAX_HISTORY_SIZE)
subscribers = []
subscribers_lock = Lock()


def _load_history_from_file():
    DATA_DIR.mkdir(exist_ok=True)

    if not HISTORY_FILE.exists():
        return

    try:
        with HISTORY_FILE.open("r", encoding="utf-8") as file:
            events = json.load(file)

        if isinstance(events, list):
            for event in events[-MAX_HISTORY_SIZE:]:
                history.append(event)

        print(f"[history] загружено {len(history)} событий из {HISTORY_FILE}")
    except Exception as error:
        print(f"[history] failed to load history: {error}")


def _save_history_to_file():
    DATA_DIR.mkdir(exist_ok=True)

    try:
        with HISTORY_FILE.open("w", encoding="utf-8") as file:
            json.dump(list(history), file, ensure_ascii=False, indent=2)
    except Exception as error:
        print(f"[history] failed to save history: {error}")


def _safe_display(value, default="unknown"):
    if not value:
        return default

    return (
        getattr(value, "display", None)
        or getattr(value, "key", None)
        or str(value)
    )


def _detect_trigger(data):
    data = data or {}

    event = str(data.get("event", "")).lower()

    if event in ("new_task", "created", "issue_created", "create"):
        return "new_task", "Новая задача"

    if event in ("status_changed", "issue_status_changed", "status"):
        return "status_changed", "Изменился статус"

    if event in ("task_updated", "updated", "issue_updated", "update"):
        return "task_updated", "Изменение задачи"

    raw = json.dumps(data, ensure_ascii=False).lower()

    if "event.create" in raw or "created" in raw:
        return "new_task", "Новая задача"

    if "status_changed" in raw:
        return "status_changed", "Изменился статус"

    return "task_updated", "Изменение задачи"


def _push_event(event):
    history.appendleft(event)
    _save_history_to_file()

    with subscribers_lock:
        dead_subscribers = []

        for subscriber in subscribers:
            try:
                subscriber.put_nowait(event)
            except Exception:
                dead_subscribers.append(subscriber)

        for subscriber in dead_subscribers:
            subscribers.remove(subscriber)


def _join_value(value):
    if value is None:
        return ""

    if isinstance(value, list):
        return " | ".join(str(item) for item in value)

    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)

    return str(value)


def record_tracker_event(issue_key, issue=None, data=None, score=None, level=None, already_processed=False):
    data = data or {}

    trigger_code, trigger_label = _detect_trigger(data)
    llm_analysis = data.get("llm_analysis")

    event = {
        "id": f"{issue_key}-{datetime.now().timestamp()}",
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "trigger_code": trigger_code,
        "trigger_label": trigger_label,
        "issue": issue_key,
        "title": getattr(issue, "summary", None) or "Без названия",
        "status": _safe_display(getattr(issue, "status", None)),
        "assignee": _safe_display(getattr(issue, "assignee", None)),
        "risk_score": round(float(score), 3) if score is not None else None,
        "risk_level": level or "unknown",
        "llm_analysis": llm_analysis,
        "already_processed": already_processed,
    }

    _push_event(event)
    return event



def _parse_history_date(value):
    if not value:
        return None

    return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()


def _event_in_period(event, start_date, end_date):
    event_date = _parse_history_date(event.get("time"))

    if event_date is None:
        return False

    if start_date and event_date < start_date:
        return False

    if end_date and event_date > end_date:
        return False

    return True


def register_frontend(app):
    _load_history_from_file()

    @app.get("/dashboard")
    def dashboard():
        return send_from_directory("frontend", "index.html")

    @app.get("/api/history")
    def api_history():
        return jsonify(list(history))


    @app.post("/api/history/clear")
    def clear_history():
        payload = request.get_json(silent=True) or {}
        mode = payload.get("mode", "all")

        if mode == "all":
            deleted_count = len(history)
            history.clear()

        elif mode == "period":
            start_date = _parse_history_date(payload.get("start_date"))
            end_date = _parse_history_date(payload.get("end_date"))

            if start_date is None and end_date is None:
                return jsonify({
                    "ok": False,
                    "error": "start_date or end_date is required",
                }), 400

            old_events = list(history)
            kept_events = []
            deleted_count = 0

            for event in old_events:
                if _event_in_period(event, start_date, end_date):
                    deleted_count += 1
                else:
                    kept_events.append(event)

            history.clear()
            history.extend(kept_events)

        else:
            return jsonify({
                "ok": False,
                "error": "unknown clear mode",
            }), 400

        _save_history_to_file()

        return jsonify({
            "ok": True,
            "deleted": deleted_count,
            "remaining": len(history),
        })


    @app.get("/api/history/export.csv")
    def export_history_csv():
        DATA_DIR.mkdir(exist_ok=True)

        export_file = DATA_DIR / "events_history_export.csv"

        fieldnames = [
            "time",
            "trigger_code",
            "issue",
            "title",
            "status",
            "assignee",
            "risk_score",
            "risk_level",
            "llm_clarity_score",
            "llm_clarity_comment",
            "llm_risk_level",
            "llm_risk_reasons",
            "llm_recommendations",
            "already_processed",
        ]

        with export_file.open("w", encoding="cp1251", errors="replace", newline="") as file:
            file.write("sep=;\n")

            writer = csv.DictWriter(
                file,
                fieldnames=fieldnames,
                delimiter=";",
                lineterminator="\n",
            )

            writer.writeheader()

            for event in history:
                llm_analysis = event.get("llm_analysis") or {}

                writer.writerow({
                    "time": event.get("time", ""),
                    "trigger_code": event.get("trigger_code", ""),
                    "issue": event.get("issue", ""),
                    "title": event.get("title", ""),
                    "status": event.get("status", ""),
                    "assignee": event.get("assignee", ""),
                    "risk_score": event.get("risk_score", ""),
                    "risk_level": event.get("risk_level", ""),
                    "llm_clarity_score": llm_analysis.get("clarity_score", ""),
                    "llm_clarity_comment": llm_analysis.get("clarity_comment", ""),
                    "llm_risk_level": llm_analysis.get("risk_level", ""),
                    "llm_risk_reasons": _join_value(llm_analysis.get("risk_reasons", "")),
                    "llm_recommendations": _join_value(llm_analysis.get("recommendations", "")),
                    "already_processed": event.get("already_processed", ""),
                })

        return send_from_directory(str(DATA_DIR), export_file.name, as_attachment=True)

    @app.get("/api/history/export.xlsx")
    def export_history_xlsx():
        DATA_DIR.mkdir(exist_ok=True)

        export_file = DATA_DIR / "events_history_export.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Events History"

        headers = [
            "Время",
            "Событие",
            "Задача",
            "Название",
            "Статус",
            "Исполнитель",
            "Risk score",
            "Уровень риска",
            "LLM: понятность",
            "LLM: комментарий",
            "LLM: риск",
            "LLM: причины риска",
            "LLM: рекомендации",
            "Уже обработано",
        ]

        trigger_labels = {
            "new_task": "Новая задача",
            "status_changed": "Изменился статус",
            "task_updated": "Изменение задачи",
        }

        risk_labels = {
            "low": "Низкий",
            "medium": "Средний",
            "high": "Высокий",
            "unknown": "Не рассчитан",
        }

        sheet.append(headers)

        for event in history:
            llm_analysis = event.get("llm_analysis") or {}

            trigger_code = event.get("trigger_code", "")
            risk_level = event.get("risk_level", "")

            sheet.append([
                event.get("time", ""),
                trigger_labels.get(trigger_code, trigger_code),
                event.get("issue", ""),
                event.get("title", ""),
                event.get("status", ""),
                event.get("assignee", ""),
                event.get("risk_score", ""),
                risk_labels.get(risk_level, risk_level),
                llm_analysis.get("clarity_score", ""),
                llm_analysis.get("clarity_comment", ""),
                llm_analysis.get("risk_level", ""),
                _join_value(llm_analysis.get("risk_reasons", "")),
                _join_value(llm_analysis.get("recommendations", "")),
                "Да" if event.get("already_processed") else "Нет",
            ])

        column_widths = {
            "A": 22,
            "B": 24,
            "C": 18,
            "D": 40,
            "E": 22,
            "F": 24,
            "G": 12,
            "H": 18,
            "I": 18,
            "J": 60,
            "K": 16,
            "L": 70,
            "M": 80,
            "N": 18,
        }

        for column, width in column_widths.items():
            sheet.column_dimensions[column].width = width

        header_fill = PatternFill("solid", fgColor="DDEBFF")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        workbook.save(export_file)

        return send_from_directory(str(DATA_DIR), export_file.name, as_attachment=True)

    @app.get("/events")
    def events():
        subscriber = Queue()

        with subscribers_lock:
            subscribers.append(subscriber)

        def stream():
            try:
                yield "retry: 3000\n\n"

                while True:
                    try:
                        event = subscriber.get(timeout=15)
                        payload = json.dumps(event, ensure_ascii=False)
                        yield f"data: {payload}\n\n"
                    except Empty:
                        yield ": keep-alive\n\n"
            finally:
                with subscribers_lock:
                    if subscriber in subscribers:
                        subscribers.remove(subscriber)

        return Response(stream(), mimetype="text/event-stream")